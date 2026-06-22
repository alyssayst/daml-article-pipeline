import csv
import io
import re
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from pydantic import BaseModel

from app.auth import CurrentUser, get_current_user
from app.batch_store import create_batch

router = APIRouter(tags=["upload"])

URL_PATTERN = re.compile(r"https?://[^\s<>\"']+")
HEADER_HINTS = {"url", "link", "href"}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100 MB


class ParseResponse(BaseModel):
    urls: list[str]
    column_name: str | None
    total_rows: int


def _looks_like_url_header(header: str) -> bool:
    return any(hint in header.lower() for hint in HEADER_HINTS)


def _extract_urls_from_rows(
    headers: list[str], rows: list[list[Any]]
) -> tuple[list[str], str | None]:
    """Scan headers and cell values to find the best URL column."""
    if not rows:
        return [], None

    # 1) Check header names first
    for idx, h in enumerate(headers):
        if _looks_like_url_header(h):
            urls = []
            for row in rows:
                if idx < len(row):
                    val = str(row[idx]).strip()
                    if URL_PATTERN.match(val):
                        urls.append(val)
            if urls:
                return urls, h

    # 2) Fall back to scanning cell values column by column
    num_cols = max(len(r) for r in rows) if rows else 0
    for idx in range(num_cols):
        urls = []
        for row in rows:
            if idx < len(row):
                val = str(row[idx]).strip()
                if URL_PATTERN.match(val):
                    urls.append(val)
        if urls and len(urls) >= len(rows) * 0.5:
            col_name = headers[idx] if idx < len(headers) else None
            return urls, col_name

    # 3) Last resort: collect all URLs found anywhere
    all_urls: list[str] = []
    for row in rows:
        for cell in row:
            for match in URL_PATTERN.findall(str(cell)):
                all_urls.append(match)
    return all_urls, None


def _parse_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []

    rows_iter = ws.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    if first_row is None:
        return [], []

    headers = [str(c) if c is not None else "" for c in first_row]
    data_rows = [[c for c in row] for row in rows_iter]
    wb.close()
    return headers, data_rows


def _parse_csv(content: bytes, delimiter: str = ",") -> tuple[list[str], list[list[Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    first_row = next(reader, None)
    if first_row is None:
        return [], []

    headers = [c.strip() for c in first_row]
    data_rows = [row for row in reader if any(c.strip() for c in row)]
    return headers, data_rows


@router.post("/upload/parse", response_model=ParseResponse)
async def parse_upload(
    file: UploadFile,
    user: CurrentUser = Depends(get_current_user),
):
    """Parse an uploaded spreadsheet and extract URLs."""
    if file.filename is None:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "csv", "tsv"):
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: .{ext}. Accepted: .xlsx, .csv, .tsv",
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 100 MB)")

    try:
        if ext == "xlsx":
            headers, data_rows = _parse_xlsx(content)
        elif ext == "tsv":
            headers, data_rows = _parse_csv(content, delimiter="\t")
        else:
            headers, data_rows = _parse_csv(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    urls, column_name = _extract_urls_from_rows(headers, data_rows)

    # Deduplicate while preserving order
    seen: set[str] = set()
    unique_urls: list[str] = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            unique_urls.append(u)

    return ParseResponse(
        urls=unique_urls,
        column_name=column_name,
        total_rows=len(data_rows),
    )


class ParseFullResponse(BaseModel):
    batch_id: str
    total_articles: int
    columns_found: list[str]


def _rows_to_article_dicts(
    headers: list[str], rows: list[list[Any]]
) -> list[dict[str, Any]]:
    """Convert header + row data into a list of dicts (one per row)."""
    articles: list[dict[str, Any]] = []
    for row in rows:
        record: dict[str, Any] = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            val = row[idx] if idx < len(row) else None
            if val is None or (isinstance(val, str) and val.strip() == ""):
                record[h] = None
            else:
                record[h] = val if not isinstance(val, str) else val.strip()
        if record:
            articles.append(record)
    return articles


@router.post("/upload/parse-full", response_model=ParseFullResponse)
async def parse_upload_full(
    file: UploadFile,
    user: CurrentUser = Depends(get_current_user),
):
    """Parse a spreadsheet and store ALL column data for filtering."""
    if file.filename is None:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "csv", "tsv"):
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: .{ext}. Accepted: .xlsx, .csv, .tsv",
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 100 MB)")

    try:
        if ext == "xlsx":
            headers, data_rows = _parse_xlsx(content)
        elif ext == "tsv":
            headers, data_rows = _parse_csv(content, delimiter="\t")
        else:
            headers, data_rows = _parse_csv(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if not data_rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    articles = _rows_to_article_dicts(headers, data_rows)

    # Deduplicate by source_url if present
    if any("source_url" in a for a in articles):
        seen: set[str] = set()
        unique: list[dict[str, Any]] = []
        for a in articles:
            url = a.get("source_url")
            if url and url in seen:
                continue
            if url:
                seen.add(url)
            unique.append(a)
        articles = unique

    batch_id = create_batch(articles, headers)

    return ParseFullResponse(
        batch_id=batch_id,
        total_articles=len(articles),
        columns_found=[h for h in headers if h],
    )
